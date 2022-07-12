from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import shutil
import time


def copy_book(name_file, path_dir, name_scheme):
    shutil.copyfile(f'{path_dir}{name_scheme}.xlsx', f'{path_dir}{name_file}.xlsx')


def read_book(i, path_dir, name_base):
    wb = load_workbook(f'{path_dir}{name_base}.xlsx')
    ws = wb.worksheets[0]
    lst = [ws["B" + str(i)].value,ws["C"+ str(i)].value,ws["D" + str(i)].value]
    return lst


def write_book(name_file, adres, podst, path_dir):
    file_name = f"{path_dir}{name_file}.xlsx"

    wb = load_workbook(filename=file_name, data_only= True)
    ws = wb.worksheets[0]

    # insert
    ws["AA54"].value = adres
    ws["AA51"].value = podst

    #center alignment
    ws["AA54"].alignment = Alignment(horizontal = 'center', vertical = 'center')
    ws["AA51"].alignment = Alignment(horizontal = 'center', vertical = 'center')

    wb.save(file_name)


def write_scheme(name_file, path_dir):
    file_name = f"{path_dir}{name_file}.xlsx"
    wb = load_workbook(file_name)
    ws = wb.worksheets[0]

    # get image
    img1 = Image(f"{path_dir}im_1.png")
    img2 = Image(f"{path_dir}im_2.png")
    img3 = Image(f"{path_dir}im_3.png")
    img4 = Image(f"{path_dir}im_3.png")

    # change of size
    img1.height = img2.height = img3.height = img4.height = 19
    img1.width = img2.width = img3.width = img4.width = 56

    # adding an image in scheme
    ws.add_image(img1, 'Y51')
    ws.add_image(img2, 'Y52')
    ws.add_image(img3, 'Y54')
    ws.add_image(img4, 'Y55')

    wb.save(file_name)


def test_time(func):
    def wrapper(*args,**kwargs):
        st = time.time()
        res = func(*args,**kwargs)
        et = time.time()
        dt = et - st
        print(f"Time work: {dt} sec")
        return res
    return wrapper


@test_time
def main(*args,**kwargs):
    for i in range(start, final):
        # read data
        data = read_book(i, path_dir, name_base)
        # get data
        name = str(data[0] + " " + data[1] + " в ст. " + data[2])
        name_accession =  name.replace('/','|')
        # copy typical scheme and rename
        copy_book(name_accession, path_dir, name_scheme)
        # write data in scheme
        adres = data[0]
        podst = data[1]+" в ст. "+data[2]
        write_book(name_accession, adres, podst, path_dir)
        # write_scheme(name_accession, path_dir)


try:
    # input info
    path_dir = input("Path to directory (/root/.../):")
    name_scheme = input("Name typical scheme:")
    name_base = input("Name base file:")
    start = int(input("Start line number base:"))
    final = int(input("Final line number base:"))

    main(path_dir, name_scheme, name_base, start, final)

except FileNotFoundError:
    print("Name directory or base file or typical scheme is incorrect")
except ValueError:
    print("Start and Final string base must be int format")
except Exception:
    print("Error")

